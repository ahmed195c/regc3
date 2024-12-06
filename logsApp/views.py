import pandas as pd
import re
from django.shortcuts import render,redirect
from django.utils import timezone
from logsApp.models import  RegistredCars , EmployesInfo,InUseCars,LogsC,FinesAccidents
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from django.db.models import Q
from django.db import IntegrityError
from .forms import UserProfileForm
import pytz

def remove_non_numeric(s):
    return re.sub(r'\D', '', s)

def removechars(c):
    characters_to_remove = "!@#$%^&*()_+|?}{:><`1234567890"

    for char in characters_to_remove:
        c = c.replace(char, "")
    
    return c
def index(request):
    return render(request, "logsApp/layout.html")

from django.shortcuts import render
from .models import InUseCars, EmployesInfo, LogsC, RegistredCars

def registerCar(request):
    all_in_use_cars = InUseCars.objects.all().order_by('-id')
    
    if request.method == "POST":
        ceo_number = request.POST.get("ceoNumber").strip()
        car_number = request.POST.get("carNumber").strip()

        try:
            its_in_use = LogsC.objects.get(Logs_employee_ins__ceoNumber=ceo_number, carIsInUse=True)
            return render(request, "logsApp/registerCar.html", {"itsinuse": its_in_use, "l": all_in_use_cars})
        except LogsC.DoesNotExist:
            pass

        try:
            car_is_in_use = LogsC.objects.get(Logs_car_ins__carNumber=car_number, carIsInUse=True)
            return render(request, "logsApp/registerCar.html", {"carIsInUse": car_is_in_use, "l": all_in_use_cars})
        except LogsC.DoesNotExist:
            pass

        try:
            emp_exists = EmployesInfo.objects.get(ceoNumber=ceo_number, EmpHaveCar=False)
        except EmployesInfo.DoesNotExist:
            emp_does_not_exist = "الرقم الاداري غير صحيح"
            return render(request, "logsApp/registerCar.html", {"empDoseNotEXISTS": emp_does_not_exist, "l": all_in_use_cars})

        try:
            car_exists = RegistredCars.objects.get(carNumber=car_number, carIsInparking=True)
        except RegistredCars.DoesNotExist:
            car_dne = "رقم المركبة غير صحيح"
            return render(request, "logsApp/registerCar.html", {"carDNE": car_dne, "l": all_in_use_cars})

        InUseCars.objects.create(car=car_exists, employee=emp_exists)
        LogsC.objects.create(Logs_employee_ins=emp_exists, Logs_car_ins=car_exists)

        emp_exists.EmpHaveCar = True
        emp_exists.save()
        car_exists.carIsInparking = False
        car_exists.save()

        success_message = "تم التسجيل بنجاح"
        return render(request, "logsApp/registerCar.html", {"sucssuMessge": success_message, "l": all_in_use_cars})

    return render(request, "logsApp/registerCar.html", {"l": all_in_use_cars})


from django.shortcuts import render
from django.utils import timezone
from .models import InUseCars, EmployesInfo, LogsC, RegistredCars
import pytz

def returnCar(request):
    if request.method == "POST":
        ceo_number = remove_non_numeric(request.POST.get("ceonumber")).strip()
        emp_note = request.POST.get("empnote")
        all_in_use_cars = InUseCars.objects.all().order_by('-id')
        dubai_tz = pytz.timezone('Asia/Dubai')

        try:
            emp_instance = EmployesInfo.objects.get(ceoNumber=ceo_number)
        except EmployesInfo.DoesNotExist:
            ret_err_msg = "الرقم الاداري غير صحيح"
            return render(request, "logsApp/registerCar.html", {"retErrm": ret_err_msg, "l": all_in_use_cars})

        try:
            in_use_car_instance = InUseCars.objects.get(employee=emp_instance)
        except InUseCars.DoesNotExist:
            ret_car_err = "لاتوجد مركبه مرتبطه بل رقم الاداري"
            return render(request, "logsApp/registerCar.html", {"retCarErr": ret_car_err, "l": all_in_use_cars})

        ret_success_msg = "تم اعاده المركبه بنجاح"
        registered_car_instance = RegistredCars.objects.get(carNumber=in_use_car_instance.car.carNumber)
        log_instance = LogsC.objects.get(Logs_employee_ins=emp_instance, carIsInUse=True)

        current_time = timezone.now().astimezone(dubai_tz)
        log_instance.ended_at = current_time
        log_instance.return_date = current_time.date()
        log_instance.return_time = current_time.time()
        log_instance.carIsInUse = False
        log_instance.carNote = emp_note

        registered_car_instance.carIsInparking = True
        emp_instance.EmpHaveCar = False

        emp_instance.save()
        log_instance.save()
        registered_car_instance.save()
        in_use_car_instance.delete()

        return render(request, "logsApp/registerCar.html", {"retSucssM": ret_success_msg, "l": all_in_use_cars})

    return render(request, "logsApp/registerCar.html", {"l": InUseCars.objects.all().order_by('-id')})



def logsfunc(request):
    if request.method == "POST":
        car_number = request.POST.get('carNumper')
        ceo_number = request.POST.get('ceoN')
        date_filter = request.POST.get('date')
        show_all = request.POST.get('showAll')

        filters = Q()

        if date_filter:
            filters &= Q(taken_date=date_filter)

        if ceo_number:
            filters &= Q(Logs_employee_ins__ceoNumber=ceo_number.strip())

        if car_number:
            filters &= Q(Logs_car_ins__carNumber=car_number.strip())

        if show_all:
            logs = LogsC.objects.all().order_by('-id')
        else:
            logs = LogsC.objects.filter(filters).order_by('-id')

        return render(request, "logsApp/logs.html", {'alllogs': logs})

    
    current_date = datetime.now().date()
    today_logs = LogsC.objects.filter(Q(taken_date=current_date) | Q(taken_date__isnull=True)).order_by('-id')
    return render(request, "logsApp/logs.html", {'alllogs': today_logs})





def finesAccidents(request):
    card1 = FinesAccidents.objects.all()
    
    if request.method == "POST":
        profile_form = UserProfileForm(request.POST, request.FILES)
        if profile_form.is_valid():
            profile_form.save()
            return redirect('logsApp:finesAccidents')
    else:
        profile_form = UserProfileForm()
    
    return render(request, "logsApp/finesaccidents.html", {
        "card1": card1,
        "profile_form": profile_form
    })

def export_to_excel(request):
    dubai_tz = pytz.timezone('Asia/Dubai')
    
    data = LogsC.objects.select_related('Logs_employee_ins', 'Logs_car_ins').all().order_by('-id')

    export_data = []
    for log in data:
        created_at_dubai = log.created_at.astimezone(dubai_tz)
        taken_time_dubai = (log.taken_time.replace(tzinfo=dubai_tz) if log.taken_time else None)
        ended_at_dubai = log.ended_at.astimezone(dubai_tz) if log.ended_at else None
        return_time_dubai = (log.return_time.replace(tzinfo=dubai_tz) if log.return_time else None)

        export_data.append({
            'ID': log.id,
            'رقم السياره': log.Logs_car_ins.carNumber,
            'الاسم': log.Logs_employee_ins.ceoName,
            'الرقم الاداري': log.Logs_employee_ins.ceoNumber,
            'تاريخ ووقت الاستلام': created_at_dubai.strftime('%Y-%m-%d %I:%M %p'),
            'تاريخ الاستلام': log.taken_date,
            'وقت الاستلام': taken_time_dubai.strftime('%I:%M %p') if taken_time_dubai else None,
            'تاريخ و وقت التسليم': ended_at_dubai.strftime('%Y-%m-%d %I:%M %p') if ended_at_dubai else None,
            'تاريخ التسليم': log.return_date,
            'وقت التسليم': return_time_dubai.strftime('%I:%M %p') if return_time_dubai else None,
            'ملاحظه على المركبه': log.carNote,
            'قسم الموظف': log.Logs_employee_ins.section
        })
    
    df = pd.DataFrame(export_data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="logs_data.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Logs Data')

        workbook = writer.book
        worksheet = writer.sheets['Logs Data']
       
        worksheet.sheet_view.rightToLeft = True
       
        header_font = Font(size=16, bold=True, color='000000')
        header_fill = PatternFill(start_color='B7E1A1', end_color='B7E1A1', fill_type='solid')
        cell_font = Font(size=16)
        center_alignment = Alignment(horizontal='center')

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
       
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.font = cell_font
                cell.alignment = center_alignment
       
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return response



def addNewEmp(request):
    if request.method == "POST":
        empNameq = request.POST.get('empName')
        empNumber = request.POST.get('empNumber')
        sa = removechars(empNameq)
        print(sa)
        try:
            EmployesInfo.objects.create(ceoName=sa,ceoNumber=empNumber)
            dd = " working"
        except IntegrityError:
            dd = "الرقم الاداري موجود من قبل"
        return render(request, "logsApp/addNewEmp.html", {'dd': dd})
    return render(request,"logsApp/addNewEmp.html")





def fineC(request):
    if request.method == "POST":
        fine_date = request.POST.get('finedate')
        fine_time = request.POST.get('finetime')
        fine_car_number = request.POST.get('finecar')
        print(f"Fine date is: {fine_date}")
        print(f"Fine time is: {fine_time}")
        dubai_tz = pytz.timezone('Asia/Dubai')
        combined_fine_datetime = dubai_tz.localize(timezone.datetime.strptime(f"{fine_date} {fine_time}", '%Y-%m-%d %H:%M'))
        print(combined_fine_datetime.time())
        try:
            car_ins = RegistredCars.objects.get(carNumber=fine_car_number)

            finon = LogsC.objects.get(Logs_car_ins=car_ins,
                                      created_at__lte = combined_fine_datetime,
                                      ended_at__gte = combined_fine_datetime
                                      )
            print(finon)
            return render(request, "logsApp/finespage.html",{"finon":finon})
        except RegistredCars.DoesNotExist:
            print(f"Car with number {fine_car_number} does not exist.")
        except Exception as e:
            print(f"An error occurred: {e}")
    return render(request,"logsApp/finespage.html",)